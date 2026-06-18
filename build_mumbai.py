import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mumbai Federer Research"

# Palette
NAVY="0D1B2A"; WHITE="FFFFFF"; LGRAY="F4F6F8"; MGRAY="DDE3EA"; DGRAY="6B7280"
G_BG="D1FAE5"; G_FG="065F46"
B_BG="DBEAFE"; B_FG="1E3A8A"
A_BG="FEF9C3"; A_FG="78350F"
R_BG="FEE2E2"; R_FG="7F1D1D"

def bd():
    s=Side(style="thin",color=MGRAY)
    return Border(left=s,right=s,top=s,bottom=s)

def paint(sheet, r, c, val="", bg=None, fg="111827", bold=False, sz=9, ha="left", wrap=True):
    cl = sheet.cell(row=r, column=c, value=val)
    if bg: cl.fill = PatternFill("solid", fgColor=bg)
    cl.font = Font(name="Calibri", bold=bold, color=fg, size=sz)
    cl.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    cl.border = bd()
    return cl

def score_col(label):
    label = (label or "").strip()
    if label in ("Strong","PASS"): return G_BG, G_FG
    if label == "Moderate":        return B_BG, B_FG
    return R_BG, R_FG

def band_col(band):
    b = band or ""
    if "A" in b: return G_BG, G_FG
    if "B" in b: return B_BG, B_FG
    if "C" in b: return A_BG, A_FG
    return R_BG, R_FG

# Column definitions: (col_idx, header, width, section)
COLS = [
    (1,"Company",20,"id"),(2,"City / Hub",13,"id"),(3,"Segment",20,"id"),
    (4,"What They Make",28,"id"),(5,"Revenue Band",13,"id"),
    (6,"Decision Maker",18,"dm"),(7,"DM Background",24,"dm"),
    (8,"E1 Producer",10,"gate"),(9,"E2 India",9,"gate"),
    (10,"C3",9,"sc"),(11,"C4",9,"sc"),(12,"C5",9,"sc"),
    (13,"C6",9,"sc"),(14,"C7",9,"sc"),(15,"C8",9,"sc"),
    (16,"Score",10,"res"),(17,"Band",18,"res"),(18,"Verdict",12,"res"),
    (19,"1-Line Reason",34,"out"),(20,"Outreach Hook",36,"out"),
]

SEC_LABEL = {
    "id":  ("COMPANY IDENTITY","1E3A5F"),
    "dm":  ("DECISION MAKER","14532D"),
    "gate":("ELIGIBILITY GATES","7F1D1D"),
    "sc":  ("FEDERER SCORE  C3–C8","3B0764"),
    "res": ("RESULT","1E3A5F"),
    "out": ("OUTREACH","14532D"),
}

# Set col widths
for col,_,w,_ in COLS:
    ws.column_dimensions[get_column_letter(col)].width = w

# Row 1: section spans
sec_ranges = {}
for col,_,_,sec in COLS:
    if sec not in sec_ranges: sec_ranges[sec]=col
for sec in ["id","dm","gate","sc","res","out"]:
    start = sec_ranges[sec]
    end = max(col for col,_,_,s in COLS if s==sec)
    label, bg = SEC_LABEL[sec]
    paint(ws,1,start,label,bg=bg,fg=WHITE,bold=True,sz=9,ha="center")
    if end > start:
        ws.merge_cells(start_row=1,start_column=start,end_row=1,end_column=end)

HDR = {"id":NAVY,"dm":"14532D","gate":"7F1D1D","sc":"3B0764","res":NAVY,"out":"14532D"}

# Row 2: column headers
for col,hdr,_,sec in COLS:
    paint(ws,2,col,hdr,bg=HDR[sec],fg=WHITE,bold=True,sz=9,ha="center")

ws.freeze_panes="A3"
ws.row_dimensions[1].height=16
ws.row_dimensions[2].height=26

# Data
records=[
    dict(company="Fineotex Chemical",city="Mumbai · Bandra",segment="Performance chemicals / textile specialty",
         what="Specialty textile chemicals, enzymes, cleaning & hygiene chemicals. 470+ SKUs, 70 countries.",
         rev="Rs.300–500Cr",dm="Surendra K. Tibrewala (CMD)\nSanjay Tibrewala (ED & CFO)",
         dm_bg="Founder 45+ yrs specialty chemicals; gen-2 CFO; external professional CEO (Arindam Choudhuri) hired",
         e1="PASS",e2="PASS",c3="Strong",c4="Strong",c5="Strong",c6="Strong",c7="Strong",c8="Strong",
         score=95,band="A — Strong Federer",verdict="Strong Pass",
         reason="45-yr operator-founder; gen-2 CFO + external CEO; 45% CAGR over 5 yrs; USD 400M revenue; 70-country distribution.",
         hook="Founder still personally receiving industry awards (April 2026) despite crossing USD 400M revenue — still actively building, not coasting."),
    dict(company="Ion Exchange (India)",city="Mumbai · Mahalaxmi",segment="Performance chemicals / water treatment",
         what="Ion exchange resins, water treatment systems, membranes, specialty chemicals for industrial water management.",
         rev="Rs.300–500Cr",dm="Dinesh Patel (CMD)",
         dm_bg="Long-tenured operator; built water treatment from niche resin chemistry; ERP-driven ops documented in annual reports",
         e1="PASS",e2="PASS",c3="Strong",c4="Strong",c5="Strong",c6="Strong",c7="Strong",c8="Moderate",
         score=88,band="A — Strong Federer",verdict="Strong Pass",
         reason="Proprietary resin IP + engineering integration moat; Jal Jeevan Mission tailwind; ERP-confirmed structured operations.",
         hook="Makes both the specialty resin AND engineers the full water treatment plant — vertical integration that competitors can't easily replicate."),
    dict(company="Rossari Biotech",city="Mumbai (HQ) · Silvassa (plant)",segment="Specialty biotech / performance chemicals",
         what="Specialty textile & home care chemicals, animal nutrition enzymes, oleo-chemicals. Contract mfg for global FMCG.",
         rev="Rs.100–300Cr",dm="Edward Menezes (MD)\nSunil Chari (MD)",
         dm_bg="Dual operator-founders; Menezes (B.Sc Chemistry, 25+ yrs); Chari (engineering); built SAP-driven multi-plant org",
         e1="PASS",e2="PASS",c3="Strong",c4="Strong",c5="Strong",c6="Strong",c7="Strong",c8="Moderate",
         score=87,band="A — Strong Federer",verdict="Strong Pass",
         reason="Dual-founder operator model; SAP-run; 3 acquisitions since IPO 2020; animal nutrition enzymes is high-growth China+1 niche.",
         hook="Made 3 structured acquisitions in 3 years post-IPO to enter animal nutrition and oleo-chemicals — M&A with clear capability logic, not financial engineering."),
    dict(company="Camlin Fine Sciences",city="Mumbai · Worli",segment="Specialty food & nutraceutical ingredients",
         what="Food antioxidants (BHA, BHT, TBHQ), catechols, specialty aroma chemicals, diphenol derivatives. 70+ country exports.",
         rev="Rs.100–300Cr",dm="Ashish Dandekar (MD & CEO)",
         dm_bg="Chemical engineer; structured international ops across multiple USFDA-approved manufacturing sites",
         e1="PASS",e2="PASS",c3="Strong",c4="Strong",c5="Strong",c6="Strong",c7="Moderate",c8="Moderate",
         score=80,band="A — Strong Federer",verdict="Pass",
         reason="Niche food antioxidants — few Indian players; USFDA-approved; clear China+1 beneficiary in a segment China still dominates.",
         hook="One of only a handful of Indian manufacturers of food-grade BHA/BHT/TBHQ — ultra-niche where China dominates and India has a real opening."),
    dict(company="Heubach Colorants India",city="Navi Mumbai",segment="Performance chemicals / pigments & dyes",
         what="Specialty pigments, pigment preparations, textile & leather dyes. For paints, plastics, printing, agrochem, personal care.",
         rev="Rs.100–300Cr",dm="Ketan Vyas (MD India)",
         dm_bg="Professional MD heading India entity; Heubach Group (Germany) parent — India entity has in-house manufacturing independence",
         e1="PASS",e2="PASS",c3="Strong",c4="Moderate",c5="Moderate",c6="Moderate",c7="Strong",c8="Moderate",
         score=72,band="B — Probable Federer",verdict="Pass",
         reason="Strong specialty pigment moat; Germany-backed process systems. Caution: verify India entity operates independently from parent group.",
         hook="Formerly Clariant Chemicals India (est. 1956) — 65+ yrs of specialty pigment manufacturing under successive global owners, local ops intact."),
    dict(company="DMCC Speciality Chemicals",city="Mumbai (HQ) · Dahej (plant)",segment="Custom synthesis / specialty chemicals",
         what="Pharma & agrochem intermediates, functional chemicals, boron chemicals. ISO 9001:2015. Est. 1919.",
         rev="Rs.100–300Cr",dm="Vivek Sharma (MD)",
         dm_bg="Chemical engineer; multi-plant ops incl. Dahej SEZ; ISO 9001:2015 certified quality management systems",
         e1="PASS",e2="PASS",c3="Strong",c4="Moderate",c5="Strong",c6="Moderate",c7="Moderate",c8="Moderate",
         score=73,band="B — Probable Federer",verdict="Pass",
         reason="105-yr specialty chemicals heritage pivoting into pharma intermediates; Dahej SEZ signals export intent. Verify current ownership structure.",
         hook="India's first-ever producer of Sulphuric Acid (1919) — now making specialty pharma intermediates from a 105-yr continuous chemistry heritage."),
    dict(company="Kilburn Engineering",city="Mumbai · Worli",segment="Industrial process equipment",
         what="Custom dryers, rotary equipment, heat exchangers for specialty chemicals, pharma, fertilizer plants.",
         rev="Rs.50–100Cr",dm="Rajeev Shukla (MD)",
         dm_bg="Engineer background; long history of custom thermal processing equipment; project mgmt systems visible in filings",
         e1="PASS",e2="PASS",c3="Moderate",c4="Moderate",c5="Strong",c6="Strong",c7="Moderate",c8="Weak",
         score=60,band="B — Probable Federer",verdict="Borderline",
         reason="Niche custom engineering moat; pharma/chemical tailwind. C8 is weak — solo MD, no visible gen-2 or professional management layer.",
         hook="Designs and builds custom dryers where every order is a unique engineering problem — a capability moat that's hard for commodity equipment players to replicate."),
    # FAILS
    dict(company="Sun Pharma",city="Mumbai",segment="Generic pharma — FAIL",
         what="Generic pharmaceuticals, branded generics, specialty drugs. India's largest pharma company.",
         rev=">Rs.500Cr",dm="Dilip Shanghvi (MD)",dm_bg="Not evaluated — auto-disqualified on revenue",
         e1="PASS",e2="PASS",c3="—",c4="—",c5="—",c6="—",c7="—",c8="—",
         score=0,band="D — Auto-disqualify",verdict="FAIL",
         reason="Revenue >Rs.14,000Cr — far exceeds ceiling. Also generic pharma (not differentiated specialty).",
         hook="—"),
    dict(company="Wockhardt",city="Mumbai · BKC",segment="Complex pharma — FAIL",
         what="Formulations, biopharmaceuticals, vaccines for global markets.",
         rev=">Rs.500Cr",dm="Murtaza Khorakiwala (MD)",dm_bg="Not evaluated — revenue + debt-restructuring disqualifies",
         e1="PASS",e2="PASS",c3="—",c4="—",c5="—",c6="—",c7="—",c8="—",
         score=0,band="D — Auto-disqualify",verdict="FAIL",
         reason="Revenue Rs.2,693Cr FY23 — exceeds ceiling. Historically debt-restructured — not a growth-mode Federer.",
         hook="—"),
    dict(company="Godrej & Boyce",city="Mumbai · Vikhroli",segment="Conglomerate — FAIL",
         what="Locks, security, aerospace systems, appliances — large diversified group division.",
         rev=">Rs.500Cr",dm="Jamshyd Godrej (CMD)",dm_bg="Not evaluated — large group subsidiary",
         e1="PASS",e2="PASS",c3="—",c4="—",c5="—",c6="—",c7="—",c8="—",
         score=0,band="D — Auto-disqualify",verdict="FAIL",
         reason="Revenue Rs.11,800Cr. Flagship of Godrej Group — fails independent promoter-driven company requirement.",
         hook="—"),
]

for ri, rec in enumerate(records):
    row = ri+3
    row_bg = WHITE if ri%2==0 else LGRAY
    vals=[rec["company"],rec["city"],rec["segment"],rec["what"],rec["rev"],
          rec["dm"],rec["dm_bg"],rec["e1"],rec["e2"],
          rec["c3"],rec["c4"],rec["c5"],rec["c6"],rec["c7"],rec["c8"],
          rec["score"],rec["band"],rec["verdict"],rec["reason"],rec["hook"]]
    for ci,(col,hdr,_,sec) in enumerate(COLS):
        val=vals[ci]; bg=row_bg; fg="111827"; bold=False; sz=9; ha="left"
        if hdr in("E1 Producer","E2 India"):
            bg,fg=score_col(str(val)); bold=True; ha="center"
        elif hdr in("C3","C4","C5","C6","C7","C8"):
            bg,fg=score_col(str(val)); bold=True; ha="center"
        elif hdr=="Score":
            s=val if isinstance(val,int) else 0
            if s>=80:  bg,fg=G_BG,G_FG
            elif s>=60:bg,fg=B_BG,B_FG
            elif s>=40:bg,fg=A_BG,A_FG
            else:      bg,fg=R_BG,R_FG
            bold=True; ha="center"; sz=12
        elif hdr=="Band":
            bg,fg=band_col(str(val)); bold=True; ha="center"
        elif hdr=="Verdict":
            v=str(val)
            if "Pass" in v and "FAIL" not in v: bg,fg=G_BG,G_FG
            elif "Borderline" in v:             bg,fg=A_BG,A_FG
            else:                               bg,fg=R_BG,R_FG
            bold=True; ha="center"
        elif hdr=="Company":
            bold=True; sz=10
        elif hdr in("Revenue Band","City / Hub"):
            ha="center"
        paint(ws,row,col,val,bg=bg,fg=fg,bold=bold,sz=sz,ha=ha)
    ws.row_dimensions[row].height=50

# Legend
ls=wb.create_sheet("Legend")
legend=[("Band","Score","Meaning","Priority"),
        ("A — Strong Federer","80–100","High-priority. Include with confidence.","Pursue"),
        ("B — Probable Federer","60–79","Good fit. Include with minor caveats.","Include"),
        ("C — Borderline","40–59","Research further before committing.","Dig deeper"),
        ("D — Disqualify","< 40","Wrong profile — too big, PE-owned, trader, or no systems.","Document & skip")]
bstyles=[(NAVY,WHITE),(G_BG,G_FG),(B_BG,B_FG),(A_BG,A_FG),(R_BG,R_FG)]
for ri,(rd,(rbg,rfg)) in enumerate(zip(legend,bstyles)):
    for ci,v in enumerate(rd):
        paint(ls,ri+2,ci+1,v,bg=rbg,fg=rfg,bold=(ri==0),sz=10,ha="center" if ci>0 else "left")
for c,wd in zip([1,2,3,4],[22,14,38,18]):
    ls.column_dimensions[get_column_letter(c)].width=wd
for r in range(2,7): ls.row_dimensions[r].height=24

crit=[("Criterion","Wt","Strong","Moderate","Weak"),
      ("C3 Differentiated","20","Patents/DSIR/USFDA/EU-GMP; proprietary products; specialized equipment","Some tech depth but no formal IP","Commodity; no moat"),
      ("C4 DM Quality","15","PhD/IIT/IISc OR operator who built ERP/costing/planning","Gen-2 with formal education OR some structured evidence","Non-technical; outsource mindset"),
      ("C5 Growing Sector","15","PLI eligible; China+1; Make-in-India; export tailwinds","Stable sector; no tailwinds","Declining or stagnant"),
      ("C6 Growth Signals","15","2+ of: Hiring/Facility/Cert/Active website/Revenue growth","1 signal","No activity in 2 yrs"),
      ("C7 Systems Maturity","20","SAP/ERP confirmed; structured costing; MIS dashboards","Some IT evidence but depth unclear","No ERP; founder intuition only"),
      ("C8 Succession","15","Gen-2 on board (operational) + external professional managers","Gen-2 exists but unclear; OR 1 professional hire","Solo founder; no gen-2")]
for ri,rd in enumerate(crit):
    bg=NAVY if ri==0 else (WHITE if ri%2==0 else LGRAY)
    fg=WHITE if ri==0 else "111827"
    for ci,v in enumerate(rd):
        paint(ls,ri+9,ci+1,v,bg=bg,fg=fg,bold=(ri==0),sz=9)
for c,wd in zip([1,2,3,4,5],[20,6,36,30,24]):
    ls.column_dimensions[get_column_letter(c)].width=wd
for r in range(9,16): ls.row_dimensions[r].height=36

out="/mnt/user-data/outputs/DeepThought_Mumbai_Federer.xlsx"
wb.save(out)
print("Saved:",out)
